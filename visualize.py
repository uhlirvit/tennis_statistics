"""
visualize.py — Standalone chart generator for TK Sport Kolovraty statistics.

Reads directly from generated xlsx files in output/, so it works independently
of main.py and reflects any manual edits you've made to those files.

Usage:
    python3 visualize.py                            # all xlsx in output/
    python3 visualize.py output/2026_dospeli.xlsx   # specific file(s)

Generates three PNG charts per team and for the all-club combined view,
saved to output/charts/{filename_stem}/{scope}_{chart}.png

Charts produced:
  1_vazeno.png     — Body Váženo horizontal bar chart
  2_cumulative.png — Cumulative Body Váženo round-by-round
  3_performance.png — Odehráno celkem vs Body Váženo (commitment vs performance)
  4_specialist.png — Singl vs Debl win-rate butterfly chart

Requires: pip install matplotlib openpyxl
"""

import sys
import os
import argparse
from pathlib import Path
from collections import defaultdict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patheffects as pe
import matplotlib.ticker as mticker
import numpy as np
from openpyxl import load_workbook

# ── Colours (same palette as the Excel data bars) ─────────────────────────
C_BLUE    = "#008AEF"
C_STEELBL = "#638EC6"
C_GREEN   = "#63C384"
C_ORANGE  = "#FFB628"
C_MAGENTA = "#D6007B"

# Distinct team colours — stable order regardless of how many teams
TEAM_COLORS = {
    "A": "#1A6FBF",
    "B": "#E87722",
    "C": "#2A9D3F",
    "D": "#8E44AD",
}

def _team_color(label):
    """Return a colour for a team label, falling back to a default palette."""
    if label in TEAM_COLORS:
        return TEAM_COLORS[label]
    extra = ["#C0392B", "#16A085", "#D4AC0D", "#884EA0", "#2C3E50"]
    return extra[hash(label) % len(extra)]

BG   = "#FAFAFA"
GRID = "#E8E8E8"

def _style():
    plt.rcParams.update({
        "font.family":          "DejaVu Sans",
        "font.size":            10,
        "axes.spines.top":      False,
        "axes.spines.right":    False,
        "axes.spines.left":     False,
        "axes.spines.bottom":   False,
        "axes.grid":            True,
        "axes.grid.axis":       "x",
        "grid.color":           GRID,
        "grid.linewidth":       0.8,
        "figure.facecolor":     BG,
        "axes.facecolor":       BG,
        "xtick.bottom":         False,
        "ytick.left":           False,
    })


# ── Read data from xlsx ────────────────────────────────────────────────────

def _headers(ws, row_idx=1):
    """Map column header value → 0-based column index."""
    return {cell.value: cell.column - 1
            for cell in ws[row_idx] if cell.value is not None}


def read_player_summary(ws) -> list[dict]:
    h = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[h.get("Jméno", 1)]
        if not name:
            continue
        rows.append({
            "name":     name,
            "teams":    row[h.get("Týmy", 0)] or "",       # only in All Club sheet
            "os":       row[h.get("Odehráno singl", 2)] or 0,
            "od":       row[h.get("Odehráno debl", 3)] or 0,
            "celkem":   row[h.get("Odehráno celkem", 4)] or 0,
            "vs":       row[h.get("Vyhráno singl", 5)] or 0,
            "vd":       row[h.get("Vyhráno debl", 6)] or 0,
            "us_singl": row[h.get("Úspěšnost singl", 7)] or 0.0,
            "us_debl":  row[h.get("Úspěšnost debl", 8)] or 0.0,
            "suma":     row[h.get("Body Suma", 9)] or 0,
            "vazeno":   row[h.get("Body Váženo", 10)] or 0.0,
        })
    return rows


def read_match_log(ws) -> list[dict]:
    h = _headers(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[h.get("Jméno", 5)]:
            continue
        rows.append({
            "kolo":    row[h.get("Kolo", 0)],
            "player":  row[h.get("Jméno", 5)],
            "type":    row[h.get("Typ", 4)],
            "result":  row[h.get("Výsledek", 7)],
        })
    return rows


def extract_workbook(xlsx_path: str) -> dict:
    """
    Parse a season workbook.  Returns:
    {
      team_label: {"summary": [...], "log": [...], "color": "#hex"},
      "All Club": {"summary": [...], "log": [...], "color": "#hex"}  # if present
    }

    Handles both prefixed sheets ("A - Player Summary") and unprefixed
    sheets ("Player Summary") — the latter occur when team_label is blank,
    e.g. for dorost or baby categories with a single unnamed team.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    teams = {}

    for sheet_name in wb.sheetnames:
        # Prefixed: "A - Player Summary", "B - Match Log", etc.
        if " - Player Summary" in sheet_name:
            label = sheet_name.replace(" - Player Summary", "")
            teams.setdefault(label, {})
            teams[label]["summary"] = read_player_summary(wb[sheet_name])
            teams[label]["color"]   = _team_color(label)

        elif " - Match Log" in sheet_name:
            label = sheet_name.replace(" - Match Log", "")
            teams.setdefault(label, {})
            teams[label]["log"] = read_match_log(wb[sheet_name])

        # Unprefixed: single team with no label (dorost, baby, etc.)
        elif sheet_name == "Player Summary":
            teams.setdefault("", {})
            teams[""]["summary"] = read_player_summary(wb[sheet_name])
            teams[""]["color"]   = C_BLUE

        elif sheet_name == "Match Log":
            teams.setdefault("", {})
            teams[""]["log"] = read_match_log(wb[sheet_name])

    # All-club sheets
    if "All Club - Player Summary" in wb.sheetnames:
        all_log = []
        for label, data in teams.items():
            for row in data.get("log", []):
                all_log.append(dict(row, team=label))
        teams["All Club"] = {
            "summary": read_player_summary(wb["All Club - Player Summary"]),
            "log":     all_log,
            "color":   C_BLUE,
        }

    return teams


# ── Chart 1: Body Váženo bar ───────────────────────────────────────────────

def chart_vazeno(summary, title, color=C_BLUE, meta=""):
    active = [s for s in summary if s["celkem"] > 0]
    if not active:
        return None
    active.sort(key=lambda s: s["vazeno"])

    names  = [s["name"] for s in active]
    values = [s["vazeno"] for s in active]
    n = len(active)

    fig, ax = plt.subplots(figsize=(9, max(4.5, n * 0.45)), facecolor=BG)
    ax.set_facecolor(BG)

    cmap   = plt.cm.Blues
    norms  = np.linspace(0.32, 0.82, n)
    colors = [cmap(v) for v in norms]

    bars = ax.barh(names, values, color=colors, height=0.65, zorder=3)
    for bar, val in zip(bars, values):
        if val > 0:
            ax.text(val + max(values) * 0.01, bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}", va="center", ha="left",
                    fontsize=9, fontweight="bold", color="#333333")

    ax.set_xlabel("Body Váženo", labelpad=8, color="#555555")
    ax.set_title(title, fontsize=13, fontweight="bold", pad=14, color="#222222")
    if meta:
        ax.text(0.99, 0.01, meta, transform=ax.transAxes,
                ha="right", va="bottom", fontsize=8, color="#BBBBBB")
    ax.set_xlim(0, max(values) * 1.18 if values else 1)
    ax.tick_params(axis="y", labelsize=9.5)
    ax.tick_params(axis="x", labelsize=8, colors="#888888")
    ax.grid(axis="x", color=GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    fig.tight_layout(pad=1.2)
    return fig


# ── Chart 2: Cumulative round-by-round ────────────────────────────────────

def _deconflict(positions: list[float], min_gap: float) -> list[float]:
    """
    Push label y-positions apart so none are closer than min_gap.
    Greedy top-down pass then bottom-up correction.
    """
    idx  = sorted(range(len(positions)), key=lambda i: -positions[i])
    out  = list(positions)

    # Top-down: each label must be at least min_gap below the previous
    prev = float("inf")
    for i in idx:
        out[i] = min(out[i], prev - min_gap)
        prev = out[i]

    # Bottom-up: push labels back up if they drifted too low
    prev = float("-inf")
    for i in reversed(idx):
        out[i] = max(out[i], prev + min_gap)
        prev = out[i]

    return out


def chart_cumulative(log, title, color=C_BLUE, meta=""):
    if not log:
        return None

    # Build per-player per-round Váženo
    pr = defaultdict(lambda: defaultdict(float))
    for row in log:
        try:
            kolo = int(row["kolo"])
        except (TypeError, ValueError):
            continue
        pts = (1.0 if row["type"] == "singl" else 0.5) if row["result"] == "W" else 0.0
        pr[row["player"]][kolo] += pts

    all_rounds = sorted({k for rd in pr.values() for k in rd})
    if not all_rounds:
        return None

    # Compute cumulative series
    series = {}
    for name, rd in pr.items():
        cum, total = [], 0.0
        for k in all_rounds:
            total += rd.get(k, 0)
            cum.append(total)
        series[name] = cum

    # Rank by final score
    final = {n: s[-1] for n, s in series.items()}
    ordered = sorted(series, key=lambda n: final[n], reverse=True)

    # Top 9 get opaque lines + labels; rest are faded
    TOP = min(9, len(ordered))
    top_set = set(ordered[:TOP])

    palette = plt.cm.tab20.colors

    fig, ax = plt.subplots(figsize=(11, 6.5), facecolor=BG)
    ax.set_facecolor(BG)
    ax.grid(axis="y", color=GRID, linewidth=0.8, zorder=0)
    ax.grid(axis="x", color=GRID, linewidth=0.4, zorder=0)

    for i, name in enumerate(ordered):
        c      = palette[i % len(palette)]
        is_top = name in top_set
        ax.plot(all_rounds, series[name],
                color=c, lw=2.0 if is_top else 0.8,
                alpha=1.0 if is_top else 0.28,
                marker="o", markersize=4 if is_top else 2,
                zorder=4 if is_top else 2)

    # Label deconfliction for top players
    top_ordered = [n for n in ordered if n in top_set]
    raw_y  = [series[n][-1] for n in top_ordered]
    colors = [palette[ordered.index(n) % len(palette)] for n in top_ordered]

    y_range  = max(max(v[-1] for v in series.values()), 0.1)
    min_gap  = y_range * 0.055
    adj_y    = _deconflict(raw_y, min_gap)

    x_end = all_rounds[-1]
    for name, ay, ry, c in zip(top_ordered, adj_y, raw_y, colors):
        # Connector dot if label moved
        if abs(ay - ry) > min_gap * 0.3:
            ax.plot([x_end, x_end + 0.15], [ry, ay], color=c,
                    lw=0.7, alpha=0.5, zorder=5)
        ax.text(x_end + 0.25, ay, name, va="center", ha="left",
                fontsize=8.5, color=c, fontweight="bold",
                path_effects=[pe.withStroke(linewidth=2.5, foreground=BG)])

    ax.set_xticks(all_rounds)
    ax.set_xticklabels([f"Kolo {k}" for k in all_rounds], fontsize=9)
    ax.set_xlim(all_rounds[0] - 0.4, all_rounds[-1] + 3.8)
    ax.set_ylabel("Body Váženo (kumulativně)", labelpad=8, color="#555555")
    ax.set_title(title, fontsize=13, fontweight="bold", pad=14, color="#222222")
    if meta:
        ax.text(0.99, 0.01, meta, transform=ax.transAxes,
                ha="right", va="bottom", fontsize=8, color="#BBBBBB")
    ax.tick_params(colors="#888888")
    fig.tight_layout(pad=1.2)
    return fig


# ── Chart 3: Commitment vs Performance scatter ─────────────────────────────

def chart_performance(summary, title, color=C_BLUE, meta="", scope_label=""):
    """
    X: Odehráno celkem (commitment / availability)
    Y: Body Váženo (performance / contribution)
    Median reference lines divide the chart into four natural quadrants.
    NOTE: path_effects are intentionally NOT used here — they caused a
    segfault in some matplotlib builds on Linux.
    """
    active = [s for s in summary if s.get("celkem", 0) > 0]
    if not active:
        return None

    xs     = [s["celkem"]  for s in active]
    ys     = [s["vazeno"]  for s in active]
    names  = [s["name"]    for s in active]
    teams  = [str(s.get("teams", scope_label) or scope_label) for s in active]

    unique_teams = sorted(set(teams))
    if len(unique_teams) == 1:
        dot_colors = [color] * len(active)
    else:
        dot_colors = [_team_color(t.split(",")[0].strip()) if t else C_STEELBL
                      for t in teams]

    fig, ax = plt.subplots(figsize=(9, 7.5), facecolor=BG)
    ax.set_facecolor(BG)

    ax.scatter(xs, ys, c=dot_colors, s=100, alpha=0.85, zorder=4,
               edgecolors="white", linewidths=0.8)

    # Safe range — guard against all-identical values (e.g. test data)
    x_range = max(max(xs) - min(xs), 1)
    y_range = max(max(ys) - min(ys), 0.5)
    med_x   = sorted(xs)[len(xs) // 2]
    med_y   = sorted(ys)[len(ys) // 2]

    ax.axvline(med_x, color="#CCCCCC", lw=1.0, ls="--", zorder=1)
    ax.axhline(med_y, color="#CCCCCC", lw=1.0, ls="--", zorder=1)

    # Quadrant corner labels
    pad_x, pad_y = x_range * 0.04, y_range * 0.04
    for qx, qy, ql, ha, va in [
        (max(xs) - pad_x, max(ys) - pad_y, "Páteř týmu",   "right", "top"),
        (min(xs) + pad_x, max(ys) - pad_y, "Specialisté",  "left",  "top"),
        (min(xs) + pad_x, min(ys) + pad_y, "Rezervy",      "left",  "bottom"),
        (max(xs) - pad_x, min(ys) + pad_y, "Dřevorubci",   "right", "bottom"),
    ]:
        ax.text(qx, qy, ql, ha=ha, va=va, fontsize=8,
                color="#CCCCCC", fontstyle="italic", zorder=1)

    # Player name labels — simple bbox instead of path_effects
    for x, y, name, c in zip(xs, ys, names, dot_colors):
        parts = name.split()
        short = f"{parts[-1]} {parts[0][0]}." if len(parts) > 1 else name
        ax.annotate(
            short, (x, y), xytext=(4, 4), textcoords="offset points",
            fontsize=8, color="#333333",
            bbox=dict(boxstyle="round,pad=0.1", fc="white", ec="none", alpha=0.55),
        )

    # Team colour legend (multi-team only)
    if len(unique_teams) > 1:
        from matplotlib.lines import Line2D
        handles = [Line2D([0], [0], marker="o", color="w",
                          markerfacecolor=_team_color(t.split(",")[0].strip()),
                          markersize=9, label=f"Tým {t}")
                   for t in unique_teams]
        ax.legend(handles=handles, loc="lower right", framealpha=0.85, fontsize=9)

    ax.set_xlabel("Odehráno celkem (zápasů)", labelpad=8, color="#555555")
    ax.set_ylabel("Body Váženo",               labelpad=8, color="#555555")
    ax.set_title(title, fontsize=13, fontweight="bold", pad=14, color="#222222")
    if meta:
        ax.text(0.99, 0.01, meta, transform=ax.transAxes,
                ha="right", va="bottom", fontsize=8, color="#BBBBBB")

    ax.set_xlim(min(xs) - x_range * 0.12, max(xs) + x_range * 0.18)
    ax.set_ylim(min(ys) - y_range * 0.15, max(ys) + y_range * 0.12)
    ax.tick_params(colors="#888888")
    ax.grid(axis="both", color=GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    fig.tight_layout(pad=1.2)
    return fig


# ── Chart 4: Singl vs Debl specialist butterfly ────────────────────────────

def chart_specialist(summary, title, meta=""):
    """
    Butterfly (diverging) bar chart — each player gets two bars:
      ← green bar (left)   = singl win rate
      → orange bar (right) = debl win rate

    Players sorted by (singl% − debl%) descending so singl specialists
    rise to the top and debl specialists sink to the bottom.

    Bars use the actual win rate (vs/os) regardless of the ≥5-match
    display threshold, but bars for players with <5 matches in a
    discipline are shown at 40% opacity to signal limited sample.

    Only players who played at least one match in BOTH disciplines are
    shown — it doesn't make sense to call someone a 'specialist' when
    they haven't been tested in the other.
    """
    active = [s for s in summary
              if s.get("os", 0) >= 1 and s.get("od", 0) >= 1
              and s.get("celkem", 0) > 0]
    if not active:
        return None

    # Compute true win rates (not capped by the ≥5 display threshold)
    for p in active:
        p["_us_s"]  = p["vs"] / p["os"] if p["os"] else 0.0
        p["_us_d"]  = p["vd"] / p["od"] if p["od"] else 0.0
        p["_delta"] = p["_us_s"] - p["_us_d"]

    # Sort: singl specialists at the top, debl at the bottom
    active.sort(key=lambda p: p["_delta"], reverse=True)

    names  = [p["name"] for p in active]
    us_s   = [p["_us_s"]  for p in active]
    us_d   = [p["_us_d"]  for p in active]
    few_s  = [p["os"] < 5 for p in active]
    few_d  = [p["od"] < 5 for p in active]
    deltas = [p["_delta"]  for p in active]

    n       = len(active)
    fig_h   = max(5, n * 0.46)
    fig, ax = plt.subplots(figsize=(12, fig_h), facecolor=BG)
    ax.set_facecolor(BG)

    y = np.arange(n)

    for i in range(n):
        s, d = us_s[i], us_d[i]
        fs, fd = few_s[i], few_d[i]

        # Singl bar — extends LEFT (negative x)
        a_s = 0.38 if fs else 0.85
        ax.barh(y[i], -s, height=0.62, left=0,
                color=C_GREEN, alpha=a_s, zorder=3)

        # Debl bar — extends RIGHT (positive x)
        a_d = 0.38 if fd else 0.85
        ax.barh(y[i], d, height=0.62, left=0,
                color=C_ORANGE, alpha=a_d, zorder=3)

        # Value labels
        label_kw = dict(va="center", fontsize=8.5, zorder=6)
        if s > 0.04:
            ax.text(-s - 0.015, y[i], f"{s:.0%}",
                    ha="right", color="#1E6B1E" if not fs else "#AAAAAA",
                    fontweight="bold" if not fs else "normal", **label_kw)
        if d > 0.04:
            ax.text(d + 0.015, y[i], f"{d:.0%}",
                    ha="left", color="#8B5000" if not fd else "#AAAAAA",
                    fontweight="bold" if not fd else "normal", **label_kw)

    # ---- Center: player name labels at x = 0 with a white background ----
    for i, name in enumerate(names):
        ax.text(0, y[i], f"  {name}  ", va="center", ha="center",
                fontsize=9, color="#222222", zorder=7,
                bbox=dict(boxstyle="square,pad=0.15", fc=BG, ec="none"))

    # ---- Role badge on the far right ─────────────────────────────────────
    for i, (delta, p) in enumerate(zip(deltas, active)):
        if p["os"] >= 5 and p["od"] >= 5:
            if delta >= 0.25:
                badge, color = "singl ▲", "#1E6B1E"
            elif delta <= -0.25:
                badge, color = "debl ▲", "#8B5000"
            elif p["_us_s"] >= 0.65 and p["_us_d"] >= 0.65:
                badge, color = "univerzál", "#444444"
            else:
                badge, color = "", "#CCCCCC"
        else:
            badge, color = "", "#DDDDDD"

        if badge:
            ax.text(1.08, y[i], badge, va="center", ha="left",
                    fontsize=8, color=color,
                    transform=ax.get_yaxis_transform())

    # ---- Zone backgrounds ────────────────────────────────────────────────
    # Thin colored gradient behind the bars to reinforce direction
    ax.axvspan(-1.0, 0, alpha=0.025, color=C_GREEN,  zorder=0)
    ax.axvspan( 0,  1.0, alpha=0.025, color=C_ORANGE, zorder=0)

    # ---- Axes & cosmetics ────────────────────────────────────────────────
    ax.axvline(0, color="#888888", lw=1.3, zorder=4)

    # Column headers above the bars
    top = n - 0.3
    ax.text(-0.5, top + 0.8, "← SINGL", ha="center", va="bottom",
            fontsize=10, color=C_GREEN, fontweight="bold")
    ax.text( 0.5, top + 0.8, "DEBL →", ha="center", va="bottom",
            fontsize=10, color=C_ORANGE, fontweight="bold")

    # Tick labels on x-axis: show % as positive regardless of sign
    ax.xaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f"{abs(v):.0%}"))
    ax.set_xlim(-1.12, 1.12)
    ax.set_ylim(-0.8, n + 0.9)

    # Hide the default y-tick labels (we draw our own name labels)
    ax.set_yticks(y)
    ax.set_yticklabels([""] * n)

    ax.set_xlabel(
        "Úspěšnost (%)  —  šedě = méně než 5 odehraných zápasů v disciplíně",
        labelpad=8, color="#666666", fontsize=9)
    ax.set_title(title, fontsize=13, fontweight="bold", pad=18, color="#222222")
    if meta:
        ax.text(0.99, -0.04, meta, transform=ax.transAxes,
                ha="right", va="top", fontsize=8, color="#BBBBBB")

    ax.grid(axis="x", color=GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.tick_params(colors="#888888")

    fig.tight_layout(pad=1.4)
    return fig

def _save(fig, path):
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=BG)
    plt.close(fig)
    print(f"  → {path}")


# ── Process one workbook ───────────────────────────────────────────────────

def process_workbook(xlsx_path: str, charts_base: str):
    stem = Path(xlsx_path).stem          # e.g. "2026_dospeli"
    # Infer season/category from filename for titling
    parts = stem.split("_", 1)
    season   = parts[0] if parts else stem
    category = parts[1] if len(parts) > 1 else ""
    meta = f"TK Sport Kolovraty  ·  {season}  ·  {category}"

    print(f"\nReading {xlsx_path}")
    teams = extract_workbook(xlsx_path)
    if not teams:
        print("  No recognisable sheets found — skipping.")
        return

    # Output folder: output/charts/2026_dospeli/
    out_dir = os.path.join(charts_base, stem)
    os.makedirs(out_dir, exist_ok=True)

    for label, data in teams.items():
        summary = data.get("summary", [])
        log     = data.get("log", [])
        color   = data.get("color", C_BLUE)

        scope     = ('all_club' if label == 'All Club'
                   else ('tym_' + label if label else 'tym'))
        title_pfx = (f'Klub — {season} / {category}' if label == 'All Club'
                     else (f'Tým {label} — {season} / {category}' if label
                           else f'{season} / {category}'))

        scope_label = label if label not in ("All Club", "") else ""

        display = label or '(bez označení)'
        print(f'  [{display}]', end='')

        # 1 — Váženo bar
        fig = chart_vazeno(summary, f'Body Váženo  ·  {title_pfx}', color, meta)
        if fig:
            _save(fig, os.path.join(out_dir, f'{scope}_1_vazeno.png'))

        # 2 — Cumulative
        fig = chart_cumulative(log, f'Vývoj bodů  ·  {title_pfx}', color, meta)
        if fig:
            _save(fig, os.path.join(out_dir, f'{scope}_2_cumulative.png'))

        # 3 — Commitment vs performance scatter
        fig = chart_performance(summary, f'Zápasy vs Body Váženo  ·  {title_pfx}',
                                color, meta, scope_label)
        if fig:
            _save(fig, os.path.join(out_dir, f'{scope}_3_performance.png'))

        # 4 — Specialist butterfly
        # For the all-club view the win rates are COMBINED across every team
        # a player plays for — so Péťa's singl % here is his total singl
        # record across Tým A + Tým C together, not per-team.  Per-team
        # charts (_tym_A_4_specialist.png etc.) show per-team rates and
        # will match the per-team Ocenění sheets exactly.
        specialist_title = f'Singlista vs Deblista  ·  {title_pfx}'
        if label == 'All Club':
            specialist_title += '\n(hodnoty kombinovány přes všechny týmy)'
        fig = chart_specialist(summary, specialist_title, meta)
        if fig:
            _save(fig, os.path.join(out_dir, f'{scope}_4_specialist.png'))



# ── Entry point ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate seasonal charts from TK Sport Kolovraty xlsx files.")
    parser.add_argument("files", nargs="*",
                        help="xlsx files to process. Default: all xlsx in output/")
    parser.add_argument("--output-dir", default="output",
                        help="Directory containing the xlsx files (default: output)")
    args = parser.parse_args()

    _style()

    if args.files:
        paths = args.files
    else:
        out = args.output_dir
        paths = sorted(str(p) for p in Path(out).glob("*.xlsx")
                       if p.name != "dryrun_test.xlsx"
                       and "history" not in p.name)
        if not paths:
            print(f"No xlsx files found in {out}/")
            return

    charts_base = os.path.join(args.output_dir, "charts")

    for path in paths:
        process_workbook(path, charts_base)

    print(f"\nDone. Charts saved under {charts_base}/")


if __name__ == "__main__":
    main()
