"""
import_history.py — ONE-TIME SCRIPT

Reads the dospeli_historie_2025, tab from the original Excel workbook
and produces history_store.json, which becomes the persistent accumulator
going forward. Only needs to be re-run if you want to rebuild from the
Excel source; normally the JSON is the source of truth and is updated
by main.py each time a season is scraped.

Run once:
    python3 import_history.py

Edit EXCEL_PATH below to point to your Excel file.
"""

import json, warnings
from pathlib import Path
import openpyxl

EXCEL_PATH    = "Statistika_mistraky.xlsx"
OUTPUT_PATH   = "history_store.json"
HISTORY_SHEET = "dospeli_historie_2025,"   # exact name incl. trailing comma

# Known player IDs from A/C tým scraping.
# Add B tým IDs here once those teams are scraped.
KNOWN_IDS = {
    "Marian": "1020475", "Milan": "1013864", "Víťa": "1007484",
    "Béďa": "1030702",   "Péťa": "24335",    "Marlon": "1053104",
    "Kája": "1020939",   "Verča": "1006471",  "Jessica": "1045010",
    "Valča": "1050153",  "Maks": "1069295",   "Radim": "1015429",
    # B tým players -- fill in after scraping B tým:
    # "Honza Hýbl st.": "...", "Honza Hýbl ml.": "...", etc.
}

# Season-tab nickname → canonical history name.
# Only needed where they differ.
SEASON_TO_CANONICAL = {
    "Honza K.":     "Honza Königsmark",
    "Radim Jirmus": "Radim",
    "Honza Kosík":  "Radim",
    "Eric":         "Kuba Šimek",
}

def canonical(name: str) -> str:
    return SEASON_TO_CANONICAL.get(name.strip(), name.strip())


def read_aggregates(wb) -> dict:
    """
    Read the history tab aggregate totals (columns D,E,G,H,M) and team
    labels (column B) for all players (column C = canonical name).
    These are the FORMULA OUTPUTS -- correct regardless of which cells
    the formulas happen to pull from internally.
    """
    ws = wb[HISTORY_SHEET]
    agg = {}
    for r in range(5, 50):
        jméno = ws.cell(row=r, column=3).value
        if not jméno:
            continue
        os_ = ws.cell(row=r, column=4).value
        if os_ is None:
            continue
        name = jméno.strip()
        tým  = ws.cell(row=r, column=2).value
        agg[name] = {
            "teams_history": str(tým).strip() if tým else "",
            "os": int(os_),
            "od": int(ws.cell(row=r, column=5).value or 0),
            "vs": int(ws.cell(row=r, column=7).value or 0),
            "vd": int(ws.cell(row=r, column=8).value or 0),
            "seasons_total": int(ws.cell(row=r, column=13).value or 0),
        }
    return agg


def read_season(wb, sheet: str) -> dict:
    """
    Read the all-club combined section from a season tab (rows 71+,
    column O = player name). Returns {canonical_name: {os,od,vs,vd}}.
    """
    ws = wb[sheet]
    data = {}
    for r in range(71, 120):
        name = ws.cell(row=r, column=15).value
        if not name or not isinstance(name, str):
            continue
        data[canonical(name)] = {
            "os": int(ws.cell(row=r, column=17).value or 0),
            "od": int(ws.cell(row=r, column=18).value or 0),
            "vs": int(ws.cell(row=r, column=20).value or 0),
            "vd": int(ws.cell(row=r, column=21).value or 0),
        }
    return data


def build_store(aggregates: dict, season_data: dict) -> dict:
    """
    For each player in the history tab, legacy = aggregate − Σ(scraped seasons).
    This back-calculation is reliable because the aggregate formula outputs are
    correct even when the internal cell references got shuffled around during
    row reordering in the Excel.
    """
    store = {}

    for name, agg in aggregates.items():
        scraped_os = scraped_od = scraped_vs = scraped_vd = 0
        scraped_years = {}
        for year, sdata in season_data.items():
            if name in sdata:
                s = sdata[name]
                scraped_os += s["os"]; scraped_od += s["od"]
                scraped_vs += s["vs"]; scraped_vd += s["vd"]
                scraped_years[year] = s

        active_scraped = sum(
            1 for s in scraped_years.values() if s["os"] + s["od"] > 0
        )
        store[name] = {
            "player_id":     KNOWN_IDS.get(name),
            "teams_history": agg["teams_history"],
            "_legacy": {
                "os":      agg["os"] - scraped_os,
                "od":      agg["od"] - scraped_od,
                "vs":      agg["vs"] - scraped_vs,
                "vd":      agg["vd"] - scraped_vd,
                "seasons": agg["seasons_total"] - active_scraped,
            },
            "seasons": scraped_years,
        }

    # Players in season tabs but absent from history → add fresh
    for year, sdata in season_data.items():
        for name, s in sdata.items():
            if name not in store:
                store[name] = {
                    "player_id":     KNOWN_IDS.get(name),
                    "teams_history": "",
                    "_legacy": {"os": 0, "od": 0, "vs": 0, "vd": 0, "seasons": 0},
                    "seasons": {year: s},
                }
            elif year not in store[name]["seasons"]:
                store[name]["seasons"][year] = s

    return store


def add_triška(store: dict):
    """Tříška Martin: 2006-2012, 7 seasons, data provided manually."""
    per_year = {
        "2006": (6,6,6,6), "2007": (7,5,7,7), "2008": (4,4,5,3),
        "2009": (6,3,6,2), "2010": (6,4,5,5), "2011": (7,7,7,5),
        "2012": (7,4,6,4),
    }
    store["Tříška Martin"] = {
        "player_id":     None,    # fill after 2026 scrape
        "teams_history": "A",     # assumed; correct if wrong
        "_legacy": {
            "os": sum(t[0] for t in per_year.values()),   # 43
            "od": sum(t[2] for t in per_year.values()),   # 42
            "vs": sum(t[1] for t in per_year.values()),   # 33
            "vd": sum(t[3] for t in per_year.values()),   # 32
            "seasons": 7,
            "_per_year": {
                y: {"os": t[0], "vs": t[1], "od": t[2], "vd": t[3]}
                for y, t in per_year.items()
            },
        },
        "seasons": {},
    }


def verify(aggregates: dict, store: dict) -> bool:
    all_ok = True
    for name, agg in aggregates.items():
        e = store.get(name, {})
        leg = e.get("_legacy", {})
        got = leg.get("os", 0) + sum(
            s["os"] for s in e.get("seasons", {}).values()
        )
        if got != agg["os"]:
            print(f"  MISMATCH {name}: expected os={agg['os']}, got {got}")
            all_ok = False
    return all_ok


def main():
    warnings.filterwarnings("ignore")
    path = Path(EXCEL_PATH)
    if not path.exists():
        print(f"ERROR: {path.absolute()} not found. Set EXCEL_PATH correctly.")
        return

    print(f"Reading {path} ...")
    wb = openpyxl.load_workbook(str(path), data_only=True)

    aggregates = read_aggregates(wb)
    print(f"  {len(aggregates)} players in history tab.")

    s2024 = read_season(wb, "2024_dospeli")
    s2025 = read_season(wb, "2025_dospeli")
    print(f"  2024: {len(s2024)} entries  |  2025: {len(s2025)} entries")

    store = build_store(aggregates, {"2024": s2024, "2025": s2025})
    add_triška(store)

    ok = verify(aggregates, store)
    print("  Verification: ALL OK ✓" if ok else "  Verification: MISMATCHES ✗")

    out = {
        "_meta": {
            "note": (
                "Persistent history accumulator for TK Sport Kolovraty dospeli. "
                "Legacy computed as (history-tab aggregate) − (scraped season sums). "
                "Going forward, main.py appends new seasons here automatically."
            ),
            "triška_note": (
                "Tříška Martin: 2006-2012, 7 seasons, provided manually. "
                "player_id = null until 2026 is scraped. teams_history = 'A' assumed."
            ),
            "radim_note": (
                "Radim = Jirmus Radim, tracked from 2024. Pre-2024 legacy deliberately "
                "excluded in original formula (=0). No separate old-Radim entry."
            ),
            "hýbl_note": (
                "Hýbl IDs unknown until B tým is scraped. Add to KNOWN_IDS then, "
                "and to name_overrides.csv for correct display names."
            ),
        },
        "players": store,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"\nWrote {OUTPUT_PATH}  ({len(store)} players)")
    for name, e in sorted(store.items()):
        leg = e["_legacy"]
        pid = e.get("player_id") or "—"
        scount = len(e.get("seasons", {}))
        print(f"  {name.strip():<24} legacy_os={leg['os']:<4} "
              f"legacy_sez={leg['seasons']:<3} scraped_years={scount} id={pid}")


if __name__ == "__main__":
    main()
