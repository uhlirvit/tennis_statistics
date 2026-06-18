"""
End-to-end scraper for cesky-tenis.cz team competitions.

WHAT THIS DOES
  Reads teams.csv (one row per team, per season). Rows sharing the same
  (season_year, category) are grouped together and written into ONE
  workbook -- e.g. A tym and C tym, both 2025 dospeli, end up as two
  sheet-groups inside output/2025_dospeli.xlsx, rather than separate
  files. This mirrors how your own Excel kept every team for a season
  in one tab.

  For each row:
    1. Fetches the competition/team page -> roster + list of rounds
    2. Fetches each round's official match record -> per-position W/L
    3. Aggregates into a player summary (Odehráno, Vyhráno, Úspěšnost, Suma, Váženo)
  For each (season_year, category) group, once every row in it has been
  attempted:
    4. Writes one .xlsx with four sheets per team (Player Summary,
       Match Grid, Match Log, Awards), sheet names prefixed by
       team_label so multiple teams can coexist in one workbook.

HOW TO RUN
  1. Install dependencies once:
       pip install beautifulsoup4 openpyxl
  2. Open teams.csv and make sure every season/team you want is listed
     (see "ADDING A NEW SEASON OR TEAM" below).
  3. Run:
       python3 main.py
  4. One file per (season_year, category) group appears in /output,
     named {season_year}_{category}.xlsx. Compare against your
     existing sheet for that season.

ADDING A NEW SEASON OR TEAM
  Add a row to teams.csv. Five things are needed, and only one of them
  comes from the website:

  - url -- the team's page on the site, e.g.
    https://cesky-tenis.cz/soutez/9029?druzstvoCislo=8
    competition_id and team_number are parsed straight out of this
    automatically, so there's nothing else to copy from the address bar.
  - season_year -- just a label, e.g. 2024. Rows that share season_year
    AND category end up in the same output file.
  - category -- dospeli / dorost / zactvo / etc. Defaults to "dospeli"
    if left blank. This exists mainly so a future youth category gets
    its own file rather than mixing into the adult one.
  - team_label -- A / B / C / etc. Used as the sheet-name prefix.
  - competition_name -- free text, purely for your own readability
    when skimming the CSV.

  Find the team's standings page on the site the same way as before
  (club page -> team -> that season), copy the URL out of the address
  bar, paste it in. If it's easier, just paste the URL in chat and it
  can be added as a row directly.

  The parser itself needs no changes for a new season *in the expected
  case*, since the page structure is config-independent -- but this is
  an expectation, not a guarantee, which is exactly why testing a new
  season or team is worth doing before trusting it long-term. One real
  example already found this way: lower teams can be listed under a
  suffixed name (e.g. "TK Sport Kolovraty C") rather than the bare club
  name -- handled automatically now, but it's the kind of thing that
  only surfaces by actually running a new row.

NOTE ON ROBOTS.TXT
  cesky-tenis.cz's robots.txt disallows automated tools, which is why
  this must be run from your own machine rather than by Claude directly.
  This script is intentionally slow (2s delay between requests) and
  only reads public results -- the same pages you'd open by hand in a
  browser. Running N rows fetches roughly 9*N pages in total.
"""

import os
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

from fetch import fetch
from parser import parse_roster, parse_rounds, parse_match_record, parse_club_display_name
from aggregate import (
    build_long_log, compute_player_summary, order_by_roster,
    compute_oceneni, AWARD_CATEGORIES, build_wide_grid, compute_all_club_summary,
)

# ---------------------------------------------------------------- CONFIG
BASE_URL = "https://cesky-tenis.cz"
CLUB_NAME = "TK Sport Kolovraty"
TEAMS_CONFIG_PATH = "teams.csv"
OUTPUT_DIR = "output"
DEFAULT_CATEGORY = "dospeli"
# --------------------------------------------------------------------


def team_url(competition_id, team_number):
    return f"{BASE_URL}/soutez/{competition_id}?druzstvoCislo={team_number}"


def match_url(competition_id, zapas_id):
    return f"{BASE_URL}/oficialni-zapis/{competition_id}?zapas={zapas_id}"


def parse_team_url(url):
    """
    Pulls competition_id and team_number straight out of a team URL like
    https://cesky-tenis.cz/soutez/9029?druzstvoCislo=8
    so teams.csv only needs to store the URL itself -- one source of
    truth instead of the URL plus two numbers that could drift out of
    sync with it.
    """
    comp_match = re.search(r"/soutez/(\d+)", url)
    team_match = re.search(r"druzstvoCislo=(\d+)", url)
    if not comp_match or not team_match:
        raise ValueError(f"Could not find competition_id/team_number in URL: {url!r}")
    return comp_match.group(1), team_match.group(1)


def load_teams(path):
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def group_teams(entries):
    """Group rows by (season_year, category), preserving first-appearance
    order. category defaults to DEFAULT_CATEGORY when blank/missing."""
    groups = {}
    for entry in entries:
        season_year = entry["season_year"].strip()
        category = (entry.get("category") or "").strip() or DEFAULT_CATEGORY
        groups.setdefault((season_year, category), []).append(entry)
    return groups


def process_one_team(entry):
    """
    Fetches and aggregates one team's season. Returns a dict with
    everything write_season_workbook needs for this team's sheet-group.
    Raises (RuntimeError/ValueError) on any failure -- the caller is
    expected to catch this per-team so one bad row doesn't stop the
    rest of its season group from being written.
    """
    url = entry["url"].strip()
    competition_id, team_number = parse_team_url(url)
    season_year = entry["season_year"].strip()
    team_label = entry["team_label"].strip()
    competition_name = entry.get("competition_name", "").strip()
    label = f"{season_year} {team_label} ({competition_name})"

    print(f"--- {label} ---")
    print(f"  URL: {url}")
    print(f"  Parsed competition_id={competition_id}, team_number={team_number}")
    print(f"  Fetching team page: {team_url(competition_id, team_number)}")
    comp_html = fetch(team_url(competition_id, team_number))

    club_display_name = parse_club_display_name(comp_html)
    if not club_display_name:
        print(f"  WARNING: couldn't read a team name off this page at all -- falling back to {CLUB_NAME!r}")
        club_display_name = CLUB_NAME
    elif CLUB_NAME not in club_display_name:
        print(f"  WARNING: page's team name is {club_display_name!r}, which doesn't contain "
              f"{CLUB_NAME!r} -- double check the URL/team_number for this row, this might be the wrong team.")
    elif club_display_name != CLUB_NAME:
        print(f"  Note: this team's exact name on the site is {club_display_name!r} (using that, not the bare club name)")

    roster = parse_roster(comp_html)
    rounds = parse_rounds(comp_html, club_display_name)
    print(f"  Found {len(roster)} roster entries, {len(rounds)} rounds.")

    match_records = {}
    for rnd in rounds:
        zid = rnd["zapas_id"]
        if zid is None:
            print(f"  Round {rnd['kolo']}: no zapas id found, skipping.")
            continue
        m_url = match_url(competition_id, zid)
        print(f"  Fetching round {rnd['kolo']} ({rnd['date']} vs {rnd['opponent']}): {m_url}")
        try:
            match_html = fetch(m_url)
            match_records[zid] = parse_match_record(match_html, club_display_name)
        except RuntimeError as e:
            print(f"    WARNING: {e}")

    long_log = build_long_log(rounds, match_records)
    if match_records and not long_log:
        raise RuntimeError(
            f"Fetched {len(match_records)} match record(s) but extracted 0 player "
            f"results for any of them. Most likely cause: the team's name on the "
            f"site doesn't match what was used for matching -- check whether this "
            f"division shows it under an unexpected display name."
        )
    summary = compute_player_summary(long_log)
    summary = order_by_roster(summary, roster)
    wide = build_wide_grid(rounds, long_log)

    return {
        "team_label": team_label,
        "roster": roster,
        "rounds": rounds,
        "long_log": long_log,
        "summary": summary,
        "wide": wide,
    }


def run_all():
    entries = load_teams(TEAMS_CONFIG_PATH)
    print(f"Loaded {len(entries)} row(s) from {TEAMS_CONFIG_PATH}\n")
    groups = group_teams(entries)

    summary_lines = []  # for the final printed report
    for (season_year, category), group_entries in groups.items():
        print(f"{'=' * 60}\n{season_year} / {category}\n{'=' * 60}")
        team_bundles = []
        for entry in group_entries:
            team_label = entry["team_label"].strip()
            try:
                data = process_one_team(entry)
                team_bundles.append((team_label, data))
                summary_lines.append((season_year, category, team_label, "OK", ""))
            except Exception as e:
                print(f"  FAILED: {e}")
                summary_lines.append((season_year, category, team_label, "FAILED", str(e)))
            print()

        if team_bundles:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            out_path = os.path.join(OUTPUT_DIR, f"{season_year}_{category}.xlsx")
            write_season_workbook(out_path, team_bundles)
            print(f"  Wrote: {out_path} ({len(team_bundles)} team(s): "
                  f"{', '.join(t for t, _ in team_bundles)})\n")
        else:
            print(f"  No teams succeeded for {season_year}/{category} -- nothing written.\n")

    print("=" * 60)
    print("SUMMARY")
    for season_year, category, team_label, status, detail in summary_lines:
        tag = f"{season_year}/{category}/{team_label}"
        print(f"  [{status}] {tag:<24} {detail}")


# ---------------------------------------------------------------- OUTPUT

HEADER_FILL = PatternFill("solid", start_color="DDEBF7")
HEADER_FONT = Font(bold=True)

# Exact colors pulled from the original workbook's conditional formatting
# rules (same metric -> same color in both the main table and Oceneni).
DATABAR_GREEN = "FF63C384"    # Vyhrano singl
DATABAR_ORANGE = "FFFFB628"   # Vyhrano debl
DATABAR_BLUE = "FF008AEF"     # Body Vazeno
SCALE_RED = "FFF8696B"        # Uspesnost -- low end
SCALE_YELLOW = "FFFFEB84"     # Uspesnost -- midpoint (50th percentile)
SCALE_GREEN = "FF63BE7B"      # Uspesnost -- high end

CF_DATABARS = {
    "databar_green": DATABAR_GREEN,
    "databar_orange": DATABAR_ORANGE,
    "databar_blue": DATABAR_BLUE,
}


def _sheet_name(prefix, base):
    """'A' + 'Player Summary' -> 'A - Player Summary'. Empty prefix ->
    just 'Player Summary' (used for the single-team dry-run preview)."""
    return f"{prefix} - {base}" if prefix else base


def _add_databar(ws, col_letter, first_row, last_row, color):
    if last_row < first_row:
        return
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.conditional_formatting.add(
        rng, DataBarRule(start_type="min", start_value=None, end_type="max", end_value=None, color=color)
    )


def _add_colorscale_ryg(ws, col_letter, first_row, last_row):
    if last_row < first_row:
        return
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="min", start_color=SCALE_RED,
            mid_type="percentile", mid_value=50, mid_color=SCALE_YELLOW,
            end_type="max", end_color=SCALE_GREEN,
        ),
    )


def _add_cf(ws, cf_kind, col_letter, first_row, last_row):
    if cf_kind is None:
        return
    if cf_kind == "colorscale":
        _add_colorscale_ryg(ws, col_letter, first_row, last_row)
    elif cf_kind in CF_DATABARS:
        _add_databar(ws, col_letter, first_row, last_row, CF_DATABARS[cf_kind])


def _style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def add_team_sheets(wb, team_label, roster, rounds, long_log, summary, wide):
    """Adds this one team's four sheets (Player Summary, Match Grid,
    Match Log, Awards) into an already-open workbook, with sheet names
    prefixed by team_label so multiple teams can share one workbook.
    Does not save -- caller owns the workbook lifecycle."""

    # ---- Player Summary ---------------------------------------------
    ws = wb.create_sheet(_sheet_name(team_label, "Player Summary"))
    headers = ["Č. na soupisce", "Jméno", "Odehráno singl", "Odehráno debl", "Odehráno celkem",
               "Vyhráno singl", "Vyhráno debl", "Úspěšnost singl", "Úspěšnost debl",
               "Body Suma", "Body Váženo"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, s in enumerate(summary):
        row_idx = i + 2
        ws.append([
            s["roster_no"], s["player_name"], s["odehrano_singl"], s["odehrano_debl"], s["odehrano_celkem"],
            s["vyhrano_singl"], s["vyhrano_debl"],
            s["uspesnost_singl"], s["uspesnost_debl"],
            s["body_suma"], s["body_vazeno"],
        ])
        ws.cell(row=row_idx, column=8).number_format = "0.0%"
        ws.cell(row=row_idx, column=9).number_format = "0.0%"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26
    for col in "CDEFGHIJK":
        ws.column_dimensions[col].width = 13

    last_row = 1 + len(summary)
    _add_databar(ws, "F", 2, last_row, DATABAR_GREEN)      # Vyhrano singl
    _add_databar(ws, "G", 2, last_row, DATABAR_ORANGE)     # Vyhrano debl
    _add_colorscale_ryg(ws, "H", 2, last_row)               # Uspesnost singl
    _add_colorscale_ryg(ws, "I", 2, last_row)               # Uspesnost debl
    _add_databar(ws, "K", 2, last_row, DATABAR_BLUE)        # Body Vazeno

    # ---- Match Grid (visual layout like your Excel) -----------------
    ws2 = wb.create_sheet(_sheet_name(team_label, "Match Grid"))
    headers_info = wide["headers"]
    ws2.cell(row=1, column=1, value="Jméno")
    col = 2
    for h in headers_info:
        ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws2.cell(row=1, column=col, value=f"{h['date']}\n{h['opponent']}")
        ws2.cell(row=2, column=col, value="singl")
        ws2.cell(row=2, column=col + 1, value="debl")
        col += 2
    _style_header_row(ws2, 1, col - 1)
    _style_header_row(ws2, 2, col - 1)
    ws2.row_dimensions[1].height = 40

    players_in_order = [s["player_name"] for s in summary]
    for i, name in enumerate(players_in_order):
        row_idx = i + 3
        ws2.cell(row=row_idx, column=1, value=name)
        col = 2
        for h in headers_info:
            cell_data = wide["grid"].get(name, {}).get(h["kolo"], {"singl": None, "debl": None})
            for j, kind in enumerate(("singl", "debl")):
                val = cell_data[kind]
                cell = ws2.cell(row=row_idx, column=col + j)
                if val == "W":
                    cell.value = 1
                elif val == "L":
                    cell.value = 0
                else:
                    cell.value = None
            col += 2
    ws2.column_dimensions["A"].width = 26

    # ---- Match Log (audit trail) -------------------------------------
    ws3 = wb.create_sheet(_sheet_name(team_label, "Match Log"))
    log_headers = ["Kolo", "Datum", "Soupeř", "Pozice", "Typ", "Jméno", "Partner / partnerka", "Výsledek"]
    ws3.append(log_headers)
    _style_header_row(ws3, 1, len(log_headers))
    for row in long_log:
        ws3.append([
            row["kolo"], row["date"], row["opponent"], row["position"],
            row["type"], row["player_name"], row["partner"] or "", row["result"] or "",
        ])
    for i, w in enumerate([6, 12, 32, 9, 7, 26, 26, 10]):
        ws3.column_dimensions[get_column_letter(i + 1)].width = w

    # ---- Awards / rankings (matches the "Oceneni" block) -------------
    _add_awards_sheet(wb, _sheet_name(team_label, "Awards"), summary)


def _add_awards_sheet(wb, sheet_name, summary):
    """Builds one Awards/ranking sheet from a player summary list.
    Shared by add_team_sheets (per-team) and add_all_club_sheets
    (cross-team) since the ranking logic is identical either way --
    only which summary list feeds it differs."""
    ws4 = wb.create_sheet(sheet_name)
    oceneni = compute_oceneni(summary)
    categories = list(oceneni.keys())
    percent_cats = {label for label, _field, kind, _cf in AWARD_CATEGORIES if kind == "percent"}

    ws4.cell(row=1, column=1, value="Pořadí")
    col = 2
    hodnota_col_for_category = {}
    for cat in categories:
        ws4.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws4.cell(row=1, column=col, value=cat)
        hodnota_col_for_category[cat] = col + 1
        col += 2
    _style_header_row(ws4, 1, col - 1)

    col = 2
    for cat in categories:
        ws4.cell(row=2, column=col, value="Jméno")
        ws4.cell(row=2, column=col + 1, value="Hodnota")
        col += 2
    _style_header_row(ws4, 2, col - 1)

    for i in range(len(summary)):
        row_idx = i + 3
        ws4.cell(row=row_idx, column=1, value=f"{i + 1}.")
        col = 2
        for cat in categories:
            entry = oceneni[cat][i]
            ws4.cell(row=row_idx, column=col, value=entry["name"])
            value_cell = ws4.cell(row=row_idx, column=col + 1, value=entry["value"])
            if cat in percent_cats:
                value_cell.number_format = "0.0%"
            col += 2

    ws4.column_dimensions["A"].width = 8
    col = 2
    for _ in categories:
        ws4.column_dimensions[get_column_letter(col)].width = 22
        ws4.column_dimensions[get_column_letter(col + 1)].width = 10
        col += 2

    last_row_oc = 2 + len(summary)
    cf_for_category = {label: cf for label, _field, _kind, cf in AWARD_CATEGORIES}
    for cat in categories:
        cf_kind = cf_for_category.get(cat)
        col_letter = get_column_letter(hodnota_col_for_category[cat])
        _add_cf(ws4, cf_kind, col_letter, 3, last_row_oc)


def add_all_club_sheets(wb, team_bundles):
    """
    Builds the merged 'All Club' view across every team in this
    season+category group -- mirroring the 'Klub dospělí' block at the
    bottom of the original Excel. A player who appeared on more than
    one team gets their raw counts summed across every team they
    played for (not averaged, not double-counted), with a 'Týmy'
    column showing which teams contributed.

    This intentionally re-keys by player_id rather than trusting name
    text to line up across teams -- which is exactly the thing that
    silently went wrong in the original spreadsheet (Volf Jakub was
    entered as "Kuba" on one team's block and "Kuba Volf" on another,
    and the combined table's formula ended up summing the wrong two
    rows together as a result). Computing this fresh avoids that.
    """
    combined_summary = compute_all_club_summary(team_bundles)

    ws = wb.create_sheet("All Club - Player Summary")
    headers = ["Jméno", "Týmy", "Odehráno singl", "Odehráno debl", "Odehráno celkem",
               "Vyhráno singl", "Vyhráno debl", "Úspěšnost singl", "Úspěšnost debl",
               "Body Suma", "Body Váženo"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for i, s in enumerate(combined_summary):
        row_idx = i + 2
        ws.append([
            s["player_name"], s["teams"], s["odehrano_singl"], s["odehrano_debl"], s["odehrano_celkem"],
            s["vyhrano_singl"], s["vyhrano_debl"],
            s["uspesnost_singl"], s["uspesnost_debl"],
            s["body_suma"], s["body_vazeno"],
        ])
        ws.cell(row=row_idx, column=8).number_format = "0.0%"
        ws.cell(row=row_idx, column=9).number_format = "0.0%"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    for col in "CDEFGHIJK":
        ws.column_dimensions[col].width = 13

    last_row = 1 + len(combined_summary)
    _add_databar(ws, "F", 2, last_row, DATABAR_GREEN)
    _add_databar(ws, "G", 2, last_row, DATABAR_ORANGE)
    _add_colorscale_ryg(ws, "H", 2, last_row)
    _add_colorscale_ryg(ws, "I", 2, last_row)
    _add_databar(ws, "K", 2, last_row, DATABAR_BLUE)

    _add_awards_sheet(wb, "All Club - Awards", combined_summary)


def write_season_workbook(path, team_bundles):
    """team_bundles: list of (team_label, data_dict) pairs, where
    data_dict has roster/rounds/long_log/summary/wide as returned by
    process_one_team. Creates a fresh workbook, adds every team's
    sheet-group, then the merged all-club view, and saves."""
    wb = Workbook()
    wb.remove(wb.active)
    for team_label, data in team_bundles:
        add_team_sheets(
            wb, team_label,
            data["roster"], data["rounds"], data["long_log"], data["summary"], data["wide"],
        )
    if len(team_bundles) >= 2:
        add_all_club_sheets(wb, team_bundles)
    wb.save(path)


if __name__ == "__main__":
    run_all()
